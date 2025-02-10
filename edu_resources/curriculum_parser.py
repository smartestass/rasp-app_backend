import django
import os
from openpyxl import load_workbook

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "schedule_creator.settings")
django.setup()
from edu_resources.models import Curriculum, Semester, DisciplineType, Discipline, DisciplinePart, SemesterCurrent

dir_path = 'планы'

# Список названий семестров
semesters = [f"{i} семестр" for i in
             ["Первый", "Второй", "Третий", "Четвертый", "Пятый", "Шестой", "Седьмой", "Восьмой", "Девятый", "Десятый",
              "Одиннадцатый"]]

# Список названий типов занятий
edu_types = ["Консультация", "Лекционные", "Практические", "Лабораторные"]


# Берет первый лист из файла xlsx. В таких файлах всего один лист
def take_actual_sheet(file):
    actual_sheet = load_workbook(file)[load_workbook(file).sheetnames[0]]
    return actual_sheet


def find_cell_for_semester(sheet, sem, max_row=5):
    for col in sheet.iter_cols(max_row):
        for cell in col:
            if cell.value == sem:
                print(f'{sem} - Row: {cell.row}, Col: {cell.column}')
                return cell


def find_cell_for_edu_type(sheet, semester_cell, edu_type):
    for column in sheet.iter_cols(min_row=semester_cell.row + 1,
                                  max_row=semester_cell.row + 1,
                                  min_col=semester_cell.column):
        for cell in column:
            if cell.value == edu_type:
                print(f'{edu_type} - Row: {cell.row}, Col: {cell.column}')
                return {'sem': semester_cell.value,
                        'edu_type': edu_type,
                        'cell': cell}


def find_code_header(sheet):
    target = '№'
    for col in sheet.iter_cols():
        for cell in col:
            if cell.value == target:
                print(f'{target}. Row: {cell.row}, Col: {cell.column}')
                return cell


# Находит первую ячейку с точкой в столбце с шифрами
def find_first_code(sheet, code_cell):
    for col in sheet.iter_cols(min_row=code_cell.row,
                               min_col=code_cell.column,
                               max_col=code_cell.column):
        for cell in col:
            try:
                if '.' in cell.value:
                    print(f'{cell.value}, Row: {cell.row}, Col: {cell.column}')
                    return cell
            except TypeError:
                pass


def find_discipline_header(sheet, code_header):
    target = 'Название дисциплины'
    for col in sheet.iter_cols(min_col=code_header.column,
                               min_row=code_header.row,
                               max_row=code_header.row):
        for cell in col:
            if cell.value == target:
                print(f'{target}. Row: {cell.row}, Col: {cell.column}')
                return cell


def is_true_discipline(this_row_value, next_row_value):
    try:
        next_row_value.split('.')
    except AttributeError:
        return True
    return len(this_row_value.split('.')) >= len(next_row_value.split('.'))


def find_only_true_disciplines(sheet, code_header, discipline_header, first_cell_with_code):
    cells_with_true_disciplines = []

    for col in sheet.iter_cols(min_col=code_header.column,
                               max_col=code_header.column,
                               min_row=first_cell_with_code.row):
        for cell in col:
            next_cell = sheet.cell(row=cell.row + 1, column=cell.column)
            if is_true_discipline(cell.value, next_cell.value):
                discipline_cell = sheet.cell(row=cell.row, column=discipline_header.column)
                print(f'Тек: {cell.value}, След: {next_cell.value}, Дисц: {discipline_cell.value}')
                cells_with_true_disciplines.append({'code': cell.value,
                                                    'discipline_cell': discipline_cell})
    return cells_with_true_disciplines


for file in os.listdir(dir_path):
    curriculum_object = Curriculum.objects.get_or_create(name=file)[0]
    sheet = take_actual_sheet(os.path.join(dir_path, file))

    coors_of_edu_types = []

    code_header = find_code_header(sheet)
    discipline_header = find_discipline_header(sheet, code_header)

    for sem in semesters:
        semester_cell = find_cell_for_semester(sheet, sem)
        for edu_type in edu_types:
            edu_type_coor = find_cell_for_edu_type(sheet, semester_cell, edu_type)
            if edu_type_coor:
                coors_of_edu_types.append(edu_type_coor)

    true_disciplines = find_only_true_disciplines(sheet, code_header, discipline_header,
                                                  first_cell_with_code=find_first_code(sheet, code_header))

    for d in true_disciplines:
        discipline_cell = d['discipline_cell']
        discipline_object = Discipline.objects.get_or_create(name=discipline_cell.value)[0]
        for i in coors_of_edu_types:
            hours = sheet.cell(row=discipline_cell.row, column=i['cell'].column)
            if hours.value:
                semester_object = Semester.objects.get_or_create(name=i['sem'],
                                                                 number=semesters.index(i['sem']) + 1)[0]
                semester_actual_object = SemesterCurrent.objects.get_or_create(semester=semester_object,
                                                                               curriculum=curriculum_object)[0]
                discipline_type_object = DisciplineType.objects.get_or_create(name=i['edu_type'])[0]
                discipline_part_object = DisciplinePart.objects.get_or_create(discipline=discipline_object,
                                                                              semester=semester_actual_object,
                                                                              type=discipline_type_object,
                                                                              hours=float(hours.value),
                                                                              curriculum=curriculum_object)
                print(discipline_part_object)
