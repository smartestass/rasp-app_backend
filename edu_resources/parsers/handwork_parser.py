import re
import datetime as dt
import openpyxl
import os
from collections import OrderedDict
import django
import pytz
from itertools import groupby

from django.db.models import Count

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "schedule_creator.settings")
django.setup()

from edu_resources.models import EduGroup, TeacherFromHand, LessonFromHand, Room

# os.chdir('C:\\Users\\uro8oros\\PycharmProjects\\outlookCalendar\\calendar_generator\\to_hanworked')

year = '.2025'
rasp_dir = './rasp2025/'
rasp_daily_dir = './rasp2025/daily/'
shedules = []

for file in os.listdir(rasp_dir):
    if file.endswith('.xlsx'):
        shedules.append(file)

schedules_daily = []
for file in os.listdir(rasp_daily_dir):
    if file.endswith('.xlsx'):
        schedules_daily.append(file)

# schedules_nam = []
# for file in os.listdir('./rasp2024/daily/NAM/'):
#     if file.endswith('.xlsx'):
#         schedules_nam.append(file)


teacher_template = r'[А-ЯЁ][а-яё]*\s[А-ЯЁ]\.\s*[А-ЯЁ]?\.?'
data_template = r'\(\d{1,2}.\d{1,2}[^\)]*\)'
event_dates_template = r'\d*ч.*'
event_type_template = r'[а-я]{2,}\.?'
single_data_template = r'\d{1,2}.\d{1,2}'
loop_data_template = r'\d{1,2}.\d{1,2}\s*-\s*\d{1,2}.\d{1,2}'
daily_data_template = r'\d{1,2}\D*\d*'
podgroup_template = r'\d'

moscow_tz = pytz.timezone('Europe/Moscow')

str_month_to_int = {'января': 1,
                    'февраля': 2,
                    'марта': 3,
                    'апреля': 4,
                    'мая': 5,
                    'июня': 6,
                    'июля': 7,
                    'августа': 8,
                    'сентября': 9,
                    'октября': 10,
                    'ноября': 11,
                    'декабря': 12}


def formated_dates(search_str: str, timing: dict, dates: list):
    only_single_data = search_str
    for loop_date in re.findall(loop_data_template, search_str):
        d_start, d_finish = re.split(r'\-', loop_date)

        first_ex_start = moscow_tz.localize(dt.datetime.strptime(d_start + year + '.' +
                                                                 timing['start_hours'] + '.' +
                                                                 timing['start_minutes'],
                                                                 '%d.%m.%Y.%H.%M'))
        first_ex_finish = moscow_tz.localize(dt.datetime.strptime(d_start + year + '.' +
                                                                  timing['finish_hours'] + '.' +
                                                                  timing['finish_minutes'],
                                                                  '%d.%m.%Y.%H.%M'))

        last_ex_start = moscow_tz.localize(dt.datetime.strptime(d_finish + year + '.' +
                                                                timing['start_hours'] + '.' +
                                                                timing['start_minutes'],
                                                                '%d.%m.%Y.%H.%M'))
        last_ex_finish = moscow_tz.localize(dt.datetime.strptime(d_finish + year + '.' +
                                                                 timing['finish_hours'] + '.' +
                                                                 timing['finish_minutes'],
                                                                 '%d.%m.%Y.%H.%M'))
        dates.append((first_ex_start, first_ex_finish))
        weeks = int((last_ex_start - first_ex_start).days // 7)
        counter_weeks = dt.timedelta(days=7)
        for i in range(weeks):
            dates.append((first_ex_start + counter_weeks,
                          first_ex_finish + counter_weeks))
            counter_weeks += dt.timedelta(days=7)
        only_single_data = only_single_data.replace(loop_date, '')
    for single_date in re.findall(single_data_template, only_single_data):
        ex_start = moscow_tz.localize(dt.datetime.strptime(single_date + year + '.' +
                                                           timing['start_hours'] + '.' +
                                                           timing['start_minutes'],
                                                           '%d.%m.%Y.%H.%M'))
        ex_finish = moscow_tz.localize(dt.datetime.strptime(single_date + year + '.' +
                                                            timing['finish_hours'] + '.' +
                                                            timing['finish_minutes'],
                                                            '%d.%m.%Y.%H.%M'))
        dates.append((ex_start, ex_finish))


def get_schedules_from_daily():
    lessons = []

    for file in schedules_daily:
        shedule = openpyxl.load_workbook(rasp_daily_dir + file, data_only=True).active
        print(file)
        # classes = []
        group_header = get_groupheader(shedule)

        groups = get_cells_with_groups(shedule, group_header)  # Здесь будут храниться ячейки с именами групп
        disc_list = []
        for group in groups:
            for cell in shedule.iter_rows(min_col=group.column, max_col=group.column):
                actual_cell = get_cell_with_value(cell)
                if actual_cell:
                    if re.search(teacher_template, actual_cell.value):
                        disc = get_disc_for_teacher(shedule, actual_cell, group.value, is_weekly=False)
                        disc_list.append(disc)

        for disc in disc_list:
            name_of_disc = disc[0]
            teachers = re.split(r'/', disc[1])
            types_of_event = re.split(r'/', disc[2])
            timing = disc[-2]
            rooms_raw = re.split(r'/', str(disc[-3]))
            for teacher in teachers:
                teachers_list = re.findall(teacher_template, teacher)
                index_teacher = teachers.index(teacher)

                try:
                    types_without_gaps = types_of_event[index_teacher].replace(' ', '')
                except IndexError:
                    types_without_gaps = types_of_event[0].replace(' ', '')

                try:
                    rooms = rooms_raw[index_teacher]
                except IndexError:
                    rooms = rooms_raw[0]

                podgroup_match = re.search(podgroup_template, types_without_gaps)
                podgroup = ''
                if podgroup_match:
                    podgroup = podgroup_match[0]

                type_of_event = types_without_gaps[:4].upper()
                if type_of_event == 'ЛАБО':
                    type_of_event = 'ЛАБ'
                elif type_of_event == 'ЗАЧЕ' or type_of_event == 'ЗАЧЁ':
                    type_of_event = 'ЗАЧ'
                ex_start = moscow_tz.localize(dt.datetime.strptime(
                    f"{timing['day']}.{timing['month']}.{timing['year']}.{timing['start_hours']}.{timing['start_minutes']}",
                    '%d.%m.%Y.%H.%M'))
                ex_finish = moscow_tz.localize(dt.datetime.strptime(
                    f"{timing['day']}.{timing['month']}.{timing['year']}.{timing['finish_hours']}.{timing['finish_minutes']}",
                    '%d.%m.%Y.%H.%M'))
                header = {}
                header['group'] = disc[-1]
                header['type'] = type_of_event
                header['lesson'] = name_of_disc
                if podgroup:
                    header['lesson'] += f' ({podgroup} подгруппа)'
                header['teachers'] = teachers_list
                header['day'] = ex_start.date()
                header['start'] = ex_start.time()
                header['finish'] = ex_finish.time()
                header['location'] = rooms
                header['file'] = file
                lessons.append(header)

    return lessons


# def get_schedules_from_daily_nam():
#
#     lessons = []
#
#     for file in schedules_nam:
#         shedule = openpyxl.load_workbook('./rasp2024/daily/NAM/' + file, data_only=True).active
#         print(file)
#         # classes = []
#         group_header = get_groupheader(shedule)
#
#         groups = get_cells_with_groups(shedule, group_header)  # Здесь будут храниться ячейки с именами групп
#         disc_list = []
#         for group in groups:
#             for cell in shedule.iter_rows(min_col=group.column, max_col=group.column):
#                 actual_cell = get_cell_with_value(cell)
#                 if actual_cell:
#                     if re.search(teacher_template, actual_cell.value):
#                         disc = get_disc_for_teacher_nam(shedule, actual_cell, group.value)
#                         disc_list.append(disc)
#
#         for disc in disc_list:
#             name_of_disc = disc[0]
#             teachers = re.split(r'/', disc[1])
#             types_of_event = re.split(r'/', disc[2])
#             timing = disc[-2]
#             rooms_raw = re.split(r'/', str(disc[-3]))
#             for teacher in teachers:
#                 teachers_list = re.findall(teacher_template, teacher)
#                 index_teacher = teachers.index(teacher)
#
#                 try:
#                     types_without_gaps = types_of_event[index_teacher].replace(' ', '')
#                 except IndexError:
#                     types_without_gaps = types_of_event[0].replace(' ', '')
#
#                 try:
#                     rooms = rooms_raw[index_teacher]
#                 except IndexError:
#                     rooms = rooms_raw[0]
#
#                 podgroup_match = re.search(podgroup_template, types_without_gaps)
#                 podgroup = ''
#                 if podgroup_match:
#                     podgroup = podgroup_match[0]
#
#                 type_of_event = types_without_gaps[:4].upper()
#                 if type_of_event == 'ЛАБО':
#                     type_of_event = 'ЛАБ'
#                 elif type_of_event == 'ЗАЧЕ' or type_of_event == 'ЗАЧЁ':
#                     type_of_event = 'ЗАЧ'
#                 ex_start = moscow_tz.localize(dt.datetime.strptime(
#                     f"{timing['day']}.{timing['month']}.{timing['year']}.{timing['start_hours']}.{timing['start_minutes']}",
#                                                                    '%d.%m.%Y.%H.%M'))
#                 ex_finish = moscow_tz.localize(dt.datetime.strptime(
#                     f"{timing['day']}.{timing['month']}.{timing['year']}.{timing['finish_hours']}.{timing['finish_minutes']}",
#                                          '%d.%m.%Y.%H.%M'))
#                 header = {}
#                 header['group'] = disc[-1]
#                 header['type'] = type_of_event
#                 header['lesson'] = name_of_disc
#                 if podgroup:
#                     header['lesson'] += f' ({podgroup} подгруппа)'
#                 header['teachers'] = teachers_list
#                 header['start'] = ex_start
#                 header['finish'] = ex_finish
#                 header['location'] = rooms
#                 header['file'] = file
#                 lessons.append(header)
#
#     return lessons

def get_groupheader(schedule):
    # print(file)
    group_header = {}
    while not group_header:
        for col in schedule.iter_cols():
            for cell in col:
                if cell.value == 'группа' or cell.value == 'группы':
                    # Находим номер строки и номер столбца ячейки, где встретилось это слово
                    group_header['column'] = cell.col_idx
                    group_header['row'] = cell.row
    return group_header


def get_cells_with_groups(sheet, group_header):
    groups = []  # Здесь будут храниться ячейки с именами групп

    for column in sheet.iter_cols(min_col=group_header['column'] + 1,
                                  max_col=sheet.max_column,
                                  min_row=group_header['row'],
                                  max_row=group_header['row']):
        if column[0].value:
            # print(column[0].value)
            group = column[0]
            groups.append(group)
    return groups


def get_cell_with_value(cell):
    if cell[0].value:
        cell_object = cell[0]
        return cell_object

    return False


def get_disc_for_teacher(schedule, cell, group_name, is_weekly=True):
    teachers = re.split(r'//', cell.value)
    names_of_dics = re.split(r'//', schedule[cell.row - 1][cell.col_idx - 1].value)
    types = re.split(r'//', schedule[cell.row + 1][cell.col_idx - 1].value)
    rooms = re.split(r'//', schedule[cell.row + 2][cell.col_idx - 1].value)
    counter = 1
    time_string = schedule[cell.row - counter][1].value
    while not time_string:
        counter += 1
        time_string = schedule[cell.row - counter][1].value

    time = re.split(r'\-', time_string)
    start_time, finish_time = time[0], time[1]

    timings_dict = {}

    if not is_weekly:
        # print(schedule[cell.row][0].value)
        day_string = schedule[cell.row][1].value
        counter = 0
        while not day_string:
            counter += 1
            day_string = schedule[cell.row - counter][0].value
        data_string = re.search(daily_data_template, day_string)[0]
        # print(data_string)
        day_data = data_string.split(' ')
        # print(day_data)
        timings_dict['day'] = day_data[0]
        timings_dict['month'] = str_month_to_int[day_data[1]]
        timings_dict['year'] = day_data[2]

    timings_dict['start_hours'], timings_dict['start_minutes'] = re.split(r'\.', start_time)
    timings_dict['finish_hours'], timings_dict['finish_minutes'] = re.split(r'\.', finish_time)
    disciplines = []
    if len(rooms) > 1:
        for disc in zip(names_of_dics, teachers, types, rooms):
            # print(current_group)
            new = list(disc)
            # new.append(rooms)
            new.append(timings_dict)
            new.append(group_name)
            disciplines.append(new)
            # print(new)
    else:
        for disc in zip(names_of_dics, teachers, types):
            # print(disc)
            # print(current_group)
            new = list(disc)
            new.append(rooms[0])
            new.append(timings_dict)
            new.append(group_name)
            disciplines.append(new)
            # print(new)
            # print(new)
    # print(disciplines)
    return disciplines


def get_disc_for_teacher_nam(schedule, cell, group_name):
    days = {'ПОНЕДЕЛЬНИК': (1, 4, 2024),
            'ВТОРНИК': (2, 4, 2024),
            'СРЕДА': (3, 4, 2024),
            'ЧЕТВЕРГ': (4, 4, 2024),
            'ПЯТНИЦА': (5, 4, 2024),
            'СУББОТА': (6, 4, 2024)}

    teachers = re.split(r'//', cell.value)
    names_of_dics = re.split(r'//', schedule[cell.row - 1][cell.col_idx - 1].value)
    types = re.split(r'//', schedule[cell.row + 1][cell.col_idx - 1].value)
    rooms = re.split(r'//', schedule[cell.row + 2][cell.col_idx - 1].value)
    time = re.split(r'\-', schedule[cell.row - 1][1].value)
    start_time, finish_time = time[0], time[1]

    timings_dict = {}

    day_string = schedule[cell.row][1].value
    counter = 0
    while not day_string:
        counter += 1
        day_string = schedule[cell.row - counter][0].value

    day_data_string = day_string.strip().upper()
    day_data = days.get(day_data_string)
    # print(day_data)
    timings_dict['day'] = day_data[0]
    timings_dict['month'] = day_data[1]
    timings_dict['year'] = day_data[2]

    timings_dict['start_hours'], timings_dict['start_minutes'] = re.split(r'\.', start_time)
    timings_dict['finish_hours'], timings_dict['finish_minutes'] = re.split(r'\.', finish_time)
    if len(rooms) > 1:
        for disc in zip(names_of_dics, teachers, types, rooms):
            # print(current_group)
            new = list(disc)
            # new.append(rooms)
            new.append(timings_dict)
            new.append(group_name)
            # print(new)
    else:
        for disc in zip(names_of_dics, teachers, types):
            # print(current_group)
            new = list(disc)
            new.append(rooms[0])
            new.append(timings_dict)
            new.append(group_name)
            # print(new)
    return new


def get_schedules():
    lessons = []

    for file in shedules:
        shedule = openpyxl.load_workbook(rasp_dir + file, data_only=True).active
        print(file)
        # classes = []
        group_header = get_groupheader(shedule)
        print(group_header)

        groups = get_cells_with_groups(shedule, group_header)  # Здесь будут храниться ячейки с именами групп
        print(groups)
        disc_list = []
        for group in groups:
            for cell in shedule.iter_rows(min_col=group.column,
                                          max_col=group.column):
                actual_cell = get_cell_with_value(cell)
                if actual_cell:
                    if re.search(teacher_template, actual_cell.value):
                        disc = get_disc_for_teacher(shedule, actual_cell, group.value)
                        for d in disc:
                            disc_list.append(d)

        for disc in disc_list:
            header = {}
            name_of_disc = disc[0]
            teachers = re.split(r'/', disc[1])
            types_of_event = re.split(r'/', disc[2])
            timing = disc[-2]
            rooms_raw = re.split(r'/', str(disc[-3]))
            for teacher in teachers:
                teachers_list = re.findall(teacher_template, teacher)
                index_teacher = teachers.index(teacher)
                types_dates_no_gaps = types_of_event[index_teacher].replace(' ', '')
                # print(types_dates_no_gaps)
                try:
                    rooms = rooms_raw[index_teacher]
                except IndexError:
                    rooms = rooms_raw[0]
                room_dates = re.finditer(data_template, rooms)
                room_dict = OrderedDict()
                for room_str, dates_str in dict(
                        zip(re.split(data_template, rooms), (r.group() for r in room_dates))).items():
                    # print(room_str, dates_str)
                    r_dates = []
                    formated_dates(dates_str, timing, r_dates)
                    room_dict[room_str] = r_dates
                # print(name_of_disc)
                # print(teacher)
                # print(room_dict)
                # print(rooms_raw)
                for cl in re.findall(r'\w*\([^\(\)]*\)', types_dates_no_gaps):
                    dates = []
                    if cl.find('зач') < 0:
                        # print(name_of_disc)
                        # print(teacher)
                        # print(cl)
                        type_of_event = cl[:4].upper()
                        if type_of_event == 'ЛАБО':
                            type_of_event = 'ЛАБ'
                        formated_dates(cl, timing, dates)

                    for date in dates:
                        # print(room_dict)
                        if room_dict:
                            # print(room_dict)
                            for room, dates in room_dict.items():
                                if room[:2] == ', ':
                                    room = room[2:]
                                if date in dates:
                                    header = {}
                                    header['group'] = disc[-1]
                                    header['type'] = type_of_event
                                    header['lesson'] = name_of_disc
                                    header['teachers'] = teachers_list
                                    header['day'] = date[0].date()
                                    header['start'] = date[0].time()
                                    header['finish'] = date[1].time()
                                    header['location'] = room
                                    header['file'] = file
                                    lessons.append(header)
                        else:
                            header = {}
                            header['group'] = disc[-1]
                            header['type'] = type_of_event
                            header['lesson'] = name_of_disc
                            header['teachers'] = teachers_list
                            header['day'] = date[0].date()
                            header['start'] = date[0].time()
                            header['finish'] = date[1].time()
                            header['location'] = rooms
                            header['file'] = file
                            lessons.append(header)
                        print(header)
    # for i in lessons:
    #     print(lessons)
    return lessons


def transform_data(source_data):
    result_list = []

    # Сначала группируем данные по файлу
    for file_name, file_data in groupby(sorted(source_data, key=lambda x: x['file']), key=lambda x: x['file']):
        file_info = {'file': file_name, 'groups': []}
        # Затем группируем данные по группе
        for group_name, group_data in groupby(file_data, key=lambda x: x['group']):
            group_info = {'group': group_name, 'lessons': []}
            # Создаем словари для уроков, исключая ключи 'group' и 'file'
            for lesson in group_data:
                lesson_info = {key: value for key, value in lesson.items() if key not in ['group', 'file']}
                group_info['lessons'].append(lesson_info)
            file_info['groups'].append(group_info)
        result_list.append(file_info)

    return result_list


schedules = transform_data(get_schedules())


# for i in schedules[3]:
#     print(i)


def get_or_create_teacher(name):
    obj, created = TeacherFromHand.objects.get_or_create(name=name)
    # print(obj, created)
    return obj


def get_or_create_group(name):
    obj, created = EduGroup.objects.get_or_create(name=name)
    # print(obj, created)
    return obj


def delete_data_from_old_file(file_name):
    LessonFromHand.objects.filter(file_name=file_name).delete()


def get_or_create_room(name):
    cutnames = {'проспект Нагибина, 13': 'МН13',
                'пер. Днепровский, 116, корпус 4': 'Д116-к4',
                'пер. Днепровский, 116, корпус 3': 'Д116-к3'}

    act_name = name.strip()
    for key in cutnames.keys():
        if act_name.startswith(key):
            location = key
            room_name = act_name.replace(key, cutnames[key]).strip()
            room, created = Room.objects.get_or_create(location=location,
                                                       name=room_name)
            return room
    room, created = Room.objects.get_or_create(location=act_name,
                                               name=act_name)
    return room


def fill_db(schedules):
    for schedule in schedules:
        file_name = schedule['file']
        delete_data_from_old_file(schedule['file'])
        for group in schedule['groups']:
            group_name = group['group']
            group_obj = get_or_create_group(group_name)
            for lesson in group['lessons']:
                room_obj = get_or_create_room(lesson['location'])
                # print(lesson)
                teachers = [get_or_create_teacher(teacher) for teacher in lesson['teachers']]

                lesson_obj = LessonFromHand.objects.filter(
                    teachers__in=teachers,
                    lesson_type=lesson['type'],
                    date=lesson['day'],
                    start_time=lesson['start'],
                    finish_time=lesson['finish'],
                    room=room_obj,
                    discipline=lesson['lesson']
                ).annotate(
                    num_teachers=Count("teachers", distinct=True)
                ).filter(
                    num_teachers=len(teachers)
                )

                if len(lesson_obj) > 1:
                    print(lesson)
                    print('Найдено две пары в одно время')
                    print(lesson_obj)
                elif len(lesson_obj) == 1:
                    print(lesson_obj)
                    actual_lesson = lesson_obj[0]
                    actual_lesson.room = room_obj
                    actual_lesson.groups.add(group_obj)
                    print(f'Добавлена группа {group_obj}')
                    actual_lesson.save()
                elif len(lesson_obj) == 0:
                    actual_lesson = LessonFromHand.objects.create(lesson_type=lesson['type'],
                                                                  date=lesson['day'],
                                                                  start_time=lesson['start'],
                                                                  finish_time=lesson['finish'],
                                                                  room=room_obj,
                                                                  discipline=lesson['lesson'])
                    print(f'Создана пара {lesson}')
                    actual_lesson.teachers.add(*teachers)
                    actual_lesson.groups.add(group_obj)
                    actual_lesson.save()

                # if created:
                #     obj.teachers.set(teachers)
                #     obj.save()
                #     print(f'{lesson} ДОБАВЛЕНО')
                # else:
                #     print(f"ПОВТОР {lesson}")
        print(f'{file_name} обработано')


fill_db(schedules)
