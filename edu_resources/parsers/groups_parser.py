import django
import os
from openpyxl import load_workbook
import pandas as pd

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "schedule_creator.settings")
django.setup()

from edu_resources.models import EduGroup


file_xslx = os.path.join('группы1.xlsx')

sheet = load_workbook(file_xslx)[load_workbook(file_xslx).sheetnames[0]]

NAPRAV_COL = 1
PROFILE_COL = 2
FORM_COL = 3
GROUP_COL = 4
PLAN_COL = 5


def find_value(sheet, row_number, col):
    value = sheet.cell(row=row_number, column=col).value

    while not value:
        row_number -= 1
        value = sheet.cell(row=row_number,
                            column=col).value

    # print(naprav)
    return value


df = pd.DataFrame()


for row in sheet.iter_rows(min_row=2,
                           min_col=PLAN_COL,
                           max_col=PLAN_COL):
    row_text = row[0].value
    if row_text:
        row_number = row[0].row
        for col in sheet.iter_cols(min_row=row_number,
                                   max_row=row_number,
                                   min_col=6,
                                   max_col=11):
            plan_size = col[0].value

            if plan_size:
                plan_data = {}
                plan_data['plan'] = row_text[13:22]
                plan_data['size'] = plan_size
                plan_data['course'] = sheet.cell(row=1, column=col[0].column).value
                plan_data['naprav'] = find_value(sheet, row_number, NAPRAV_COL)
                plan_data['profile'] = find_value(sheet, row_number, PROFILE_COL)
                plan_data['form'] = find_value(sheet, row_number, FORM_COL)
                plan_data['group'] = find_value(sheet, row_number, GROUP_COL)
                df = pd.concat([df, pd.DataFrame([plan_data])], ignore_index=True)


# unique_groups = df['group'].unique()
# print(len(unique_groups))
# for i in unique_groups:
#     groups_row = df[df['group'] == i]
#     print(groups_row)


total_size_per_group = df.groupby('group')['size'].sum().reset_index()

total_size_per_group.columns = ['group', 'total_size']

df['total_size'] = df['group'].map(total_size_per_group.set_index('group')['total_size'])


mask = df['size'] == df.groupby('group')['size'].transform('max')
max_rows = df[mask]
# print(max_rows)

EduGroup.objects.bulk_create([
    EduGroup(name=row['group'],
             work_plan=row['plan'],
             form=row['form'],
             naprav=row['naprav'],
             profile=row['profile'],
             size=int(row['total_size']),
             course=int(row['course'])) for _, row in max_rows.iterrows()
])

# max_rows.to_excel('try.xlsx')