import django
import os
from openpyxl import load_workbook
import pandas as pd

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "schedule_creator.settings")
django.setup()

from edu_resources.models import Teacher


file_xslx = os.path.join('Список_преподавателей.xlsx')

sheet = load_workbook(file_xslx)[load_workbook(file_xslx).sheetnames[0]]

NAME_COL = 1
KAF_COL = 2
STEP_COL = 3
ZVAN_COL = 4

df = pd.DataFrame()

for row in sheet.iter_rows(min_row=2,
                           min_col=NAME_COL,
                           max_col=ZVAN_COL):
    name = row[0].value

    if name:
        teacher_dict = {'name': name}
        teacher_dict['kaf'] = sheet.cell(row=row[0].row,
                                         column=KAF_COL).value
        teacher_dict['step'] = sheet.cell(row=row[0].row,
                                          column=STEP_COL).value
        teacher_dict['zvan'] = sheet.cell(row=row[0].row,
                                          column=ZVAN_COL).value
        df = pd.concat([df, pd.DataFrame([teacher_dict])], ignore_index=True)


Teacher.objects.bulk_create([
    Teacher(name=row['name'],
            kafedra=row['kaf'],
            stepen=row['step'],
            zvanie=row['zvan']) for _, row in df.iterrows()
])
