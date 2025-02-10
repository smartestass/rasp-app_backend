import django
import os
from openpyxl import load_workbook
import pandas as pd

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "schedule_creator.settings")
django.setup()

from edu_resources.models import Teacher, Discipline, EduGroup


file_xslx = os.path.join('Сводная учебная нагрузка ЮФУ.xls')

# sheet = load_workbook(file_xslx)[load_workbook(file_xslx).sheetnames[0]]

df = pd.read_excel(file_xslx)

work_plans_from_db = EduGroup.objects.values_list('work_plan', flat=True)

filtred_df = df[(df['Нагрузка'].isin(['Практические',
                                   'Лекционные',
                                   'Консультация',
                                   'Лабораторные',
                                   'Семинар']))
                & (df['Период'] == '2')
                & (df['Уровень подготовки'] != 'Аспирант')
                & (df['Форма обучения'].isin(['Очная', 'Очная, Очно-заочная', 'Очно-заочная']))]

mask = filtred_df['Учебный план'].apply(lambda x: any(plan in x for plan in work_plans_from_db))

doubles = []

for _, row in filtred_df[mask].iterrows():
    plans = row['Учебный план'].split(', ')
    groups = []
    for p in plans:
        try:
            group = EduGroup.objects.get(work_plan=p)
            if group:
                groups.append(group)
                print(f'Добавлена единственная группа {group}')
        except EduGroup.DoesNotExist:
            print(f'Группы c планом {p} нет в БД')
            pass
        except EduGroup.MultipleObjectsReturned:

            print(f'Для плана {p} обнаружено несколько групп')

            d = EduGroup.objects.filter(work_plan=p)
            group_names = row['Группа'].replace('Группа ', '').split(', ')

            if 'Основной поток' in group_names:
                for g in d:
                    if g not in groups:
                        groups.append(g)
                print(f'{row["Дисциплина"]} {row["Нагрузка"]} {row["Преподаватель"]}')
                print(f'Основной поток, добавлены группы {d}')

            else:
                counter = 0
                for g in d:
                    if g.name in group_names:
                        groups.append(g)
                        counter += 1
                        print(f'Добавлена группа {g.name}')
                    else:
                        print(f'Группы {g.name} нет в ячейке')
                if counter == 0:
                    print('!!!')
                    print(f'НИ ОДНА ГРУППА НЕ ДОБАВЛЕНА ДЛЯ {row["Дисциплина"]} {row["Нагрузка"]} {row["Преподаватель"]}')
                    print(f'КОД РАБОЧЕГО ПЛАНА {p} ')

    print(f'ДОБАВЛЕНИЕ ГРУПП ЗАВЕРШЕНО ДЛЯ {row["Дисциплина"]} {row["Нагрузка"]} {row["Преподаватель"]}')

    try:
        teacher = Teacher.objects.get(name=row['Преподаватель'])
    except Teacher.DoesNotExist:
        teacher = None

    disciplines = Discipline.objects.filter(name=row['Дисциплина'],
                                     semester=row['Семестр'],
                                     type=row['Нагрузка'],
                                     kafedra=row['Кафедра'],
                                     myam=row['МУАМ'])
    if teacher:
        disciplines = disciplines.filter(teachers__id=teacher.id)

    if disciplines:
        print(f'{row["Дисциплина"]} {row["Нагрузка"]} {row["Преподаватель"]} ЕСТЬ В БД')
        matching_disciplines = [
            disc for disc in disciplines if set(disc.groups.all()) == set(groups)
        ]
        if matching_disciplines:
            if 'Подгруппа' not in row['Группа']:
                discipline = matching_disciplines[0]
                print(f'{discipline} найдена. Часов - {discipline.hours}')
                discipline.hours += float(row['Аудиторные часы'].replace(',', '.'))
                print(f'Теперь часов - {discipline.hours}')
                continue

    discipline = Discipline.objects.create(name=row['Дисциплина'],
                             semester=row['Семестр'],
                             type=row['Нагрузка'],
                             kafedra=row['Кафедра'],
                             myam=row['МУАМ'],
                             hours=float(row['Аудиторные часы'].replace(',', '.')))
    discipline.groups.set(groups)
    if teacher:
        discipline.teachers.add(teacher)

    print(f'Дисциплина {discipline} добавлена')
    print('----------')

# print(len(doubles))