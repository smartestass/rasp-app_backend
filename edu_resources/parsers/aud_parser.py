import django
import os
from openpyxl import load_workbook
import pandas as pd

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "schedule_creator.settings")
django.setup()

from edu_resources.models import Room


file_xslx = os.path.join('Список_аудиторий.xlsx')

df = pd.read_excel(file_xslx)
for index, row in df.iterrows():
    if pd.isna(row['Количество мест']):
        size = None
    else:
        size = row['Количество мест']
    obj, created = Room.objects.get_or_create(name=row['Аудитория'],
                               location=row['Корпус'],
                               type=row['Тип помещения'],
                               size=size)
    if created:
        print(f'{obj.name} created')
    else:
        print(f'{obj.name} уже есть в БД')

# sheet = load_workbook(file_xslx)[load_workbook(file_xslx).sheetnames[0]]