import django
import os
import openpyxl
from openpyxl.styles import Alignment
from datetime import date, timedelta

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "schedule_creator.settings")
django.setup()
from edu_resources.models import Curriculum, Semester, DisciplineType, Discipline, DisciplinePart, SemesterCurrent, Week, EduGroup
from edu_resources.models import Lesson
from edu_resources.views import DAYS, TIME_SLOTS, fill_timetable
from edu_resources.serializers import LessonSerializer


def create_schedule_pdf(group):
    timetable = {time_slot: {day: [] for day in DAYS} for time_slot in TIME_SLOTS}
    lessons = Lesson.objects.filter(week__group=group)
    lessons_serialized = LessonSerializer(lessons, many=True).data
    fill_timetable(timetable, DAYS, lessons_serialized)
    print(timetable)
    # Создание xlsx файла
    # Создание xlsx файла
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Расписание"

    # Заполнение заголовков
    ws.append(["День недели", "Время", "Данные"])

    # Выравнивание текста в ячейках
    alignment_vertical = Alignment(text_rotation=90, vertical="top", horizontal="center")
    alignment_center = Alignment(horizontal="center", vertical="center")

    for time_slot, days in timetable.items():
        for day, lessons in days.items():
            # Если для данного дня и временного слота нет занятий, добавляем пустую строку
            if not lessons:
                ws.append([day, time_slot, ""])
                continue

            combined_data = []
            for lesson in lessons:
                discipline = f"{lesson['discipline_name']} ({lesson['discipline_type']})"
                teacher = lesson['teacher']
                dates = ", ".join([lesson['date']])
                audience = lesson.get('audience', 'не указано')
                combined_entry = f"{discipline} // {teacher} // {dates} // {audience}"
                combined_data.append(combined_entry)

            # Объединяем данные для всех дисциплин через //
            combined_cell_data = "\n\n".join(combined_data)
            ws.append([day, time_slot, combined_cell_data])

    # Форматирование первого столбца (дни недели) с вертикальной ориентацией текста
    for cell in ws["A"]:
        cell.alignment = alignment_vertical

    # Форматирование остальных ячеек с выравниванием по центру
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=3):
        for cell in row:
            cell.alignment = alignment_center

    # Сохранение файла
    wb.save("schedule.xlsx")
    print("Файл 'schedule.xlsx' успешно создан.")


group = EduGroup.objects.get(id=12)
create_schedule_pdf(group)
